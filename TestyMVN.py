import numpy as np
import numpy.linalg as alg
import scipy.linalg as salg
from scipy.stats import chi2, shapiro, norm
from sklearn.covariance import LedoitWolf
from math import exp, log, sqrt
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import load_workbook


def Mardia_test_skewness(X, standardization):
    d = X.shape[1]
    n = X.shape[0]
    X_n = np.mean(X, axis=0)
    D_n = np.zeros((n, n))
    b_nd = 0
    if standardization is True:
        S_n = np.cov(np.transpose(X), bias=True)
        S_n_inversed = alg.inv(S_n)
        for j in range(n):
            for k in range(n):
                D_n[j, k] = np.dot(np.dot(np.array([X[j, :] - X_n]), S_n_inversed), np.transpose(np.array([X[k, :] - X_n])))
    else:
        for j in range(n):
            for k in range(n):
                D_n[j, k] = np.dot(np.transpose(X[j, :]), X[k, :])
    for j in range(n):
        for k in range(n):
            b_nd += (D_n[j, k]) ** 3
    b_nd = b_nd * (1/n**2)
    return b_nd


def rozklad_empiryczny_MS():
    path = r"E:\Documents\Aartykuły PRACA MAGISTERSKA\Python\standard_True\Mardia_skew.xlsx"

    mean0 = np.zeros(2)
    I = np.identity(2)

    wyniki = []
    for j in range(10000):
        x = np.random.multivariate_normal(mean0, I, 100)
        wyniki.append(Mardia_test_skewness(x, standardization=True))
    df2 = pd.DataFrame(wyniki)

    mean0 = np.zeros(3)
    I = np.identity(3)

    wyniki3 = []
    for j in range(10000):
        x = np.random.multivariate_normal(mean0, I, 100)
        wyniki3.append(Mardia_test_skewness(x, standardization=True))
    df3 = pd.DataFrame(wyniki3)

    writer = pd.ExcelWriter(path, engine = 'openpyxl')
    df2.to_excel(writer, index=False, sheet_name = '2x100')
    df3.to_excel(writer, index=False, sheet_name = '3x100')
    writer.save()
    writer.close()

    book = load_workbook(path)
    writer = pd.ExcelWriter(path, engine = 'openpyxl')
    writer.book = book

    mean0 = np.zeros(4)
    I = np.identity(4)

    wyniki4 = []
    for j in range(10000):
        x = np.random.multivariate_normal(mean0, I, 100)
        wyniki4.append(Mardia_test_skewness(x, standardization=True))
    df4 = pd.DataFrame(wyniki4)

    mean0 = np.zeros(5)
    I = np.identity(5)

    wyniki5 = []
    for j in range(10000):
        x = np.random.multivariate_normal(mean0, I, 100)
        wyniki5.append(Mardia_test_skewness(x, standardization=True))
    df5 = pd.DataFrame(wyniki5)

    df4.to_excel(writer, index=False, sheet_name = '4x100')
    df5.to_excel(writer, index=False, sheet_name = '5x100')
    writer.save()
    writer.close()
    return 'ok'


rozklad_empiryczny_MS()


def Mardia_test_kurtosis(X, standardization):
    d = X.shape[1]
    n = X.shape[0]
    X_n = np.mean(X, axis=0)
    D_n = np.zeros((n, n))
    b_nd = 0
    if standardization is True:
        S_n = np.cov(np.transpose(X), bias=True)
        S_n_inversed = alg.inv(S_n)
        for j in range(n):
            for k in range(n):
                D_n[j, k] = np.dot(np.dot(np.array([X[j, :] - X_n]), S_n_inversed),
                                   np.transpose(np.array([X[k, :] - X_n])))
    else:
        for j in range(n):
            for k in range(n):
                D_n[j, k] = np.dot(np.transpose(X[j, :]), X[k, :])
    for j in range(n):
        b_nd += (D_n[j, j])**2
    b_nd = b_nd * (1/n)
    return b_nd


def rozklad_empiryczny_kurtoza():
    path = r"E:\Documents\Aartykuły PRACA MAGISTERSKA\Python\standard_True\Mardia_kurt.xlsx"

    mean0 = np.zeros(2)
    I = np.identity(2)

    wyniki = []
    for j in range(10000):
        x = np.random.multivariate_normal(mean0, I, 100)
        wyniki.append(Mardia_test_kurtosis(x, standardization=True))
    df2 = pd.DataFrame(wyniki)

    mean0 = np.zeros(3)
    I = np.identity(3)

    wyniki3 = []
    for j in range(10000):
        x = np.random.multivariate_normal(mean0, I, 100)
        wyniki3.append(Mardia_test_kurtosis(x, standardization=True))
    df3 = pd.DataFrame(wyniki3)

    writer = pd.ExcelWriter(path, engine = 'openpyxl')
    df2.to_excel(writer, index=False, sheet_name = '2x100')
    df3.to_excel(writer, index=False, sheet_name = '3x100')
    writer.save()
    writer.close()

    book = load_workbook(path)
    writer = pd.ExcelWriter(path, engine = 'openpyxl')
    writer.book = book

    mean0 = np.zeros(4)
    I = np.identity(4)

    wyniki4 = []
    for j in range(10000):
        x = np.random.multivariate_normal(mean0, I, 100)
        wyniki4.append(Mardia_test_kurtosis(x, standardization=True))
    df4 = pd.DataFrame(wyniki4)

    mean0 = np.zeros(5)
    I = np.identity(5)

    wyniki5 = []
    for j in range(10000):
        x = np.random.multivariate_normal(mean0, I, 100)
        wyniki5.append(Mardia_test_kurtosis(x, standardization=True))
    df5 = pd.DataFrame(wyniki5)

    df4.to_excel(writer, index=False, sheet_name = '4x100')
    df5.to_excel(writer, index=False, sheet_name = '5x100')
    writer.save()
    writer.close()
    return 'ok'


rozklad_empiryczny_kurtoza()


def HZtest(X, standardization, beta = 1):
    d = X.shape[1]
    n = X.shape[0]
    X_n = np.mean(X, axis=0)
    D_n = np.zeros((n, n))
    if standardization is True:
        S_n = np.cov(np.transpose(X), bias=True)
        S_n_inversed = alg.inv(S_n)
        for j in range(n):
            for k in range(n):
                D_n[j, k] = np.dot(np.dot(np.array([X[j, :] - X_n]), S_n_inversed),
                                   np.transpose(np.array([X[k, :] - X_n])))
    else:
        for j in range(n):
            for k in range(n):
                D_n[j, k] = np.dot(np.transpose(X[j, :]), X[k, :])
    suma1 = 0
    for j in range(n):
        for k in range(n):
            suma1 += exp((-(beta**2)/2) * (D_n[j, j] - 2*D_n[j, k] + D_n[k, k]))
    suma1 = (1/n**2) * suma1
    suma2 = 0
    for j in range(n):
        suma2 += exp((-beta**2 * D_n[j, j])/(2*(1+beta**2)))
    statystyka = suma1 - 2*(1+beta**2)**(-d/2)*(1/n)*suma2 + (1 + 2*beta**2)**(-d/2)
    return statystyka


def optimal_beta(n, d):
    h = (4/((2*d+1)*n))**(1/(d+4))
    beta = 1/(sqrt(2)*h)
    return beta


def rozklad_empiryczny_HZ():
    path = r"E:\Documents\Aartykuły PRACA MAGISTERSKA\Python\standard_True\HZtest.xlsx"

    mean0 = np.zeros(2)
    I = np.identity(2)

    wyniki = []
    for j in range(10000):
        x = np.random.multivariate_normal(mean0, I, 100)
        wyniki.append(HZtest(x, standardization=True, beta=optimal_beta(100, 2)))
    df2 = pd.DataFrame(wyniki)

    mean0 = np.zeros(3)
    I = np.identity(3)

    wyniki3 = []
    for j in range(10000):
        x = np.random.multivariate_normal(mean0, I, 100)
        wyniki3.append(HZtest(x, standardization=True, beta=optimal_beta(100, 3)))
    df3 = pd.DataFrame(wyniki3)

    mean0 = np.zeros(4)
    I = np.identity(4)

    wyniki4 = []
    for j in range(10000):
        x = np.random.multivariate_normal(mean0, I, 100)
        wyniki4.append(HZtest(x, standardization=True, beta=optimal_beta(100, 4)))
    df4 = pd.DataFrame(wyniki4)

    mean0 = np.zeros(5)
    I = np.identity(5)

    wyniki5 = []
    for j in range(10000):
        x = np.random.multivariate_normal(mean0, I, 100)
        wyniki5.append(HZtest(x, standardization=True, beta=optimal_beta(100, 5)))
    df5 = pd.DataFrame(wyniki5)

    writer = pd.ExcelWriter(path, engine = 'openpyxl')
    df2.to_excel(writer, index=False, sheet_name = '2x100')
    df3.to_excel(writer, index=False, sheet_name = '3x100')
    df4.to_excel(writer, index=False, sheet_name = '4x100')
    df5.to_excel(writer, index=False, sheet_name = '5x100')
    writer.save()
    writer.close()
    return 'ok'


rozklad_empiryczny_HZ()


def Roystest(X):
    p = X.shape[1]
    n = X.shape[0]
    W = []
    Z = []
    if n <= 3:
        print('n is too small.')
        return
    elif (n >= 4) and (n <= 11):
        x = n
        g = -2.273 + 0.459 * x
        m = 0.5440 - 0.39978 * x + 0.025054 * x**2 - 0.0006714 * x**3
        s = exp(1.3822 - 0.77857 * x + 0.062767 * x**2 - 0.0020322 * x**3)
        for j in range(0, p):
            W.append(shapiro(X[:, j])[0])
            Z.append((-log(g - (log(1 - W[j]))) - m) / s)
    elif (n >= 12) and (n <= 2000):
        x = log(n)
        g = 0
        m = -1.5861 - 0.31082 * x - 0.083751 * x**2 + 0.0038915 * x**3
        s = exp(-0.4803 - 0.082676 * x + 0.0030302 * x**2)
        for j in range(p):
            W.append(shapiro(X[:, j])[0])
            Z.append(((log(1 - W[j])) + g - m) / s)
    else:
        print('n is not in the proper size range.')
    R = []
    for j in range(p):
        R.append((norm.ppf((norm.cdf(- Z[j])) / 2))**2)
    u = 0.715
    v = 0.21364 + 0.015124 * (log(n))**2 - 0.0018034 * (log(n))**3
    l = 5
    C = np.corrcoef(X)
    NC = (C**l) * (1 - (u * (1 - C)**u) / v)
    T = np.sum(np.sum(NC)) - p
    mC = T / (p**2 - p)
    e = p / (1 + (p - 1) * mC)
    H = (e * (sum(R))) / p
    return H


def rozklad_empiryczny_Royston():
    path = r"E:\Documents\Aartykuły PRACA MAGISTERSKA\Python\Roystest.xlsx"

    mean0 = np.zeros(2)
    I = np.identity(2)

    wyniki = []
    for j in range(10000):
        x = np.random.multivariate_normal(mean0, I, 100)
        wyniki.append(Roystest(x))
    df2 = pd.DataFrame(wyniki)

    mean0 = np.zeros(3)
    I = np.identity(3)

    wyniki3 = []
    for j in range(10000):
        x = np.random.multivariate_normal(mean0, I, 100)
        wyniki3.append(Roystest(x))
    df3 = pd.DataFrame(wyniki3)

    mean0 = np.zeros(4)
    I = np.identity(4)

    wyniki4 = []
    for j in range(10000):
        x = np.random.multivariate_normal(mean0, I, 100)
        wyniki4.append(Roystest(x))
    df4 = pd.DataFrame(wyniki4)

    mean0 = np.zeros(5)
    I = np.identity(5)

    wyniki5 = []
    for j in range(10000):
        x = np.random.multivariate_normal(mean0, I, 100)
        wyniki5.append(Roystest(x))
    df5 = pd.DataFrame(wyniki5)

    writer = pd.ExcelWriter(path, engine = 'openpyxl')
    df2.to_excel(writer, index=False, sheet_name = '2x100')
    df3.to_excel(writer, index=False, sheet_name = '3x100')
    df4.to_excel(writer, index=False, sheet_name = '4x100')
    df5.to_excel(writer, index=False, sheet_name = '5x100')
    writer.save()
    writer.close()
    return "ok"


rozklad_empiryczny_Royston()